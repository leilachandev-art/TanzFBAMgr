"""
FBA Manager - Excel Data Importer
Reads your Excel file and loads all data into fba.db
Usage: python import_excel.py "path\\to\\your_file.xlsx"
"""
import sys, os
from datetime import datetime
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH    = os.path.join(SCRIPT_DIR, "backend", "fba.db")
sys.path.insert(0, os.path.join(SCRIPT_DIR, "backend"))

from sqlalchemy import create_engine, func
from sqlalchemy.orm import sessionmaker

_db_url = os.environ.get("DATABASE_URL")
if _db_url:
    if _db_url.startswith("postgres://"):
        _db_url = _db_url.replace("postgres://", "postgresql://", 1)
    engine = create_engine(_db_url, pool_size=5, max_overflow=10, pool_pre_ping=True)
    print(f"  Using cloud database (Supabase)")
else:
    engine = create_engine(f"sqlite:///{DB_PATH}", connect_args={"check_same_thread": False})
    print(f"  Using local SQLite: {DB_PATH}")

Session = sessionmaker(bind=engine)

import models
models.Base.metadata.create_all(engine)

# ── progress bar ──────────────────────────────────────────────────────────────
def progress(current, total, prefix="", width=40):
    filled = int(width * current / total) if total else 0
    bar    = "█" * filled + "░" * (width - filled)
    pct    = int(100 * current / total) if total else 0
    print(f"\r  {prefix} [{bar}] {pct}% ({current}/{total})", end="", flush=True)

def done_line(msg):
    print(f"\r  ✅ {msg}" + " " * 20)

def section(title):
    print(f"\n{'='*55}")
    print(f"  {title}")
    print(f"{'='*55}")

# ── helpers ───────────────────────────────────────────────────────────────────
def to_dt(v):
    if v is None: return None
    if isinstance(v, datetime): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%Y %H:%M %Z", "%m/%d/%Y %H:%M"):
            try: return datetime.strptime(v.split(" EDT")[0].split(" EST")[0].strip(), fmt)
            except: pass
    return None

def to_int(v, default=0):
    try: return int(v) if v not in (None, "") else default
    except: return default

def to_str(v):
    if v is None: return None
    s = str(v).strip()
    return None if not s else s

def ensure_client(db, code):
    if not code: return
    code = str(code).strip()
    if not code: return
    if not db.query(models.Client).filter(models.Client.code == code).first():
        db.add(models.Client(code=code, name=code))
        db.commit()

# ── count real rows ───────────────────────────────────────────────────────────
def count_rows(ws, start_row, max_col=30):
    total = 0
    for row in ws.iter_rows(min_row=start_row, max_col=max_col, values_only=True):
        if any(v is not None for v in row): total += 1
    return total

# ── import inbound sheet ──────────────────────────────────────────────────────
def import_main_sheet(db, ws, client_code, col_map, start_row=3):
    def val(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row): return None
        return row[idx]

    print(f"  Counting rows in sheet...", end="", flush=True)
    all_rows = [r for r in ws.iter_rows(min_row=start_row, max_col=30, values_only=True)
                if any(v is not None for v in r)]
    total = len(all_rows)
    print(f"\r  Found {total} data rows to process.")

    imported_containers = 0
    imported_outbound   = 0
    skipped_duplicate   = 0
    current_container   = None

    for i, row in enumerate(all_rows):
        progress(i + 1, total, prefix=f"  Processing (containers={imported_containers}, outbound={imported_outbound})")

        cn          = to_str(val(row, "cn"))
        client_raw  = to_str(val(row, "client")) if "client" in col_map else None
        if not client_raw: client_raw = client_code
        ctns_in     = to_int(val(row, "ctns"))
        date_in     = to_dt(val(row, "date"))

        if cn and cn not in ("CN# / JOB#", "AWB#", "CN #"):
            ensure_client(db, client_raw)
            existing = db.query(models.Container).filter(models.Container.container_no == cn).first()
            if existing:
                current_container = existing
                skipped_duplicate += 1
            else:
                current_container = models.Container(
                    client_code=str(client_raw),
                    date_in=date_in,
                    container_type=to_str(val(row, "ct")) or "container",
                    container_no=cn,
                    mark=to_str(val(row, "mark")),
                    inspection=to_str(val(row, "inspection")),
                    ctns_in=ctns_in,
                    skids_in=to_int(val(row, "skids")),
                    sku=to_str(val(row, "sku")),
                    destination=to_str(val(row, "dest")),
                    note=to_str(val(row, "note")),
                    status="In Storage",
                    created_by="import",
                )
                db.add(current_container)
                db.flush()
                imported_containers += 1

        out_date    = to_dt(val(row, "out_date"))
        out_ctns    = to_int(val(row, "out_ctns"))
        out_carrier = to_str(val(row, "out_carrier"))

        if current_container and (out_date or out_ctns or out_carrier):
            r = models.OutboundRecord(
                container_id=current_container.id,
                date_out=out_date,
                ctns_out=out_ctns,
                skids_out=to_int(val(row, "out_skids")),
                carrier=out_carrier,
                destination=to_str(val(row, "out_dest")),
                note=to_str(val(row, "out_note")),
                pod=to_str(val(row, "out_pod")),
                created_by="import",
            )
            db.add(r)
            imported_outbound += 1

        if (i + 1) % 100 == 0:
            db.commit()

    db.commit()

    print(f"\r  Updating container statuses...           ", end="", flush=True)
    for c in db.query(models.Container).all():
        total_out = db.query(func.coalesce(func.sum(models.OutboundRecord.ctns_out), 0)).filter(
            models.OutboundRecord.container_id == c.id).scalar()
        if total_out >= c.ctns_in > 0:
            c.status = "Completed"
        elif total_out > 0:
            c.status = "Partially Out"
    db.commit()

    done_line(f"Done! {imported_containers} new containers, {imported_outbound} outbound records"
              + (f", {skipped_duplicate} already existed (skipped)" if skipped_duplicate else ""))
    return imported_containers, imported_outbound

# ── import FbaCarriers sheet ──────────────────────────────────────────────────
def import_carriers(db, ws, start_row=2):
    all_rows = [r for r in ws.iter_rows(min_row=start_row, max_col=20, values_only=True)
                if any(v is not None for v in r)]
    total    = len(all_rows)
    imported = skipped = 0

    for i, row in enumerate(all_rows):
        progress(i + 1, total, prefix="  Processing")
        appt_id = to_str(row[1])
        if not appt_id: continue
        appt_id = str(appt_id)
        if db.query(models.FbaAppointment).filter(models.FbaAppointment.appointment_id == appt_id).first():
            skipped += 1; continue
        ensure_client(db, to_str(row[0]))
        db.add(models.FbaAppointment(
            client_code=to_str(row[0]), appointment_id=appt_id,
            trailer_number=to_str(row[2]), reference_code=to_str(row[3]),
            requested_delivery_date=to_str(row[4]), status=to_str(row[5]) or "PENDING",
            earliest_arrival_time=to_str(row[6]), scheduled_time=to_str(row[7]),
            arrival_time=to_str(row[8]), checkin_time=to_str(row[9]),
            unloaded_time=to_str(row[10]), closed_time=to_str(row[11]),
            creation_time=to_str(row[12]), carrier=to_str(row[13]),
            destination_fc=to_str(row[14]), isa=to_str(row[15]),
            created_by="import",
        ))
        imported += 1

    db.commit()
    done_line(f"Done! {imported} appointments imported"
              + (f", {skipped} already existed (skipped)" if skipped else ""))
    return imported

# ── import DOCKING FORM sheet ─────────────────────────────────────────────────
def import_docking(db, ws, start_row=6):
    all_rows = [r for r in ws.iter_rows(min_row=start_row, max_col=12, values_only=True)
                if any(v is not None for v in r)]
    total    = len(all_rows)
    imported = 0

    for i, row in enumerate(all_rows):
        progress(i + 1, total, prefix="  Processing")
        date_in  = to_dt(row[1])
        dock_no  = to_str(row[2])
        date_out = to_dt(row[11]) if len(row) > 11 else None

        for ctn, carrier, client, status, dout in [
            (to_str(row[3]), to_str(row[4]),  to_str(row[5]),  to_str(row[6]),  date_out),
            (to_str(row[7]), to_str(row[8]),  to_str(row[9]),  to_str(row[10]), None),
        ]:
            if not ctn or not ctn.strip("　 "): continue
            ensure_client(db, client)
            container = db.query(models.Container).filter(models.Container.container_no == ctn).first()
            db.add(models.DockingRecord(
                date_in=date_in, dock_no=dock_no,
                container_id=container.id if container else None,
                container_no=ctn, carrier=carrier, client_code=client,
                status=status or "FULL", date_out=dout, created_by="import",
            ))
            imported += 1

    db.commit()
    done_line(f"Done! {imported} docking records imported")
    return imported

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python import_excel.py \"path\\to\\file.xlsx\"")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    section("Loading workbook (may take a moment for large files)...")
    print(f"  File: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    print(f"  Sheets found: {wb.sheetnames}")

    db = Session()
    grand_containers = grand_outbound = 0

    # Inbound sheets — each has its own column layout
    INBOUND_SHEETS = {
        "2025": {
            "client": "2025",
            "col_map": {
                "client": 0, "date": 1, "ct": 2, "cn": 3, "mark": 4,
                "inspection": 5, "ctns": 6, "skids": 7, "sku": 8, "dest": 9,
                "out_date": 13, "out_ctns": 14, "out_skids": 15,
                "out_carrier": 16, "out_dest": 17, "out_note": 18, "out_pod": 19,
            },
        },
        "2033": {
            "client": "2033",
            "col_map": {
                "client": 0, "date": 1, "ct": 2, "cn": 3, "mark": 4,
                "inspection": 5, "ctns": 6, "skids": 7, "note": 8, "sku": 9, "dest": 10,
                "out_date": 14, "out_ctns": 15, "out_skids": 16,
                "out_carrier": 17, "out_dest": 18, "out_note": 19,
            },
        },
        "Voltz": {
            "client": "Voltz",
            "col_map": {
                "date": 0, "ct": 1, "cn": 2, "inspection": 3,
                "ctns": 4, "skids": 5, "mark": 6, "sku": 7, "dest": 8,
                "out_date": 12, "out_ctns": 13, "out_skids": 14,
                "out_carrier": 15, "out_dest": 16, "out_note": 17,
            },
        },
        "3002-241230": {
            "client": "3002",
            "col_map": {
                "date": 0, "cn": 1, "inspection": 2, "ctns": 3, "skids": 4,
                "out_date": 7, "out_ctns": 8, "out_skids": 9,
                "out_carrier": 10, "out_dest": 11, "out_note": 12,
            },
        },
        "3014-HMMU6739805": {
            "client": "3014",
            "col_map": {
                "date": 0, "cn": 1, "inspection": 2, "ctns": 3, "skids": 4,
                "out_date": 7, "out_ctns": 8, "out_skids": 9,
                "out_carrier": 10, "out_dest": 11, "out_note": 12,
            },
        },
        "3008-24001": {
            "client": "3008",
            "col_map": {
                "date": 0, "cn": 1, "inspection": 2, "ctns": 3, "skids": 4,
                "out_date": 7, "out_ctns": 8, "out_skids": 9,
                "out_carrier": 10, "out_dest": 11, "out_note": 12,
            },
        },
    }

    for sheet_name, cfg in INBOUND_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        section(f"Importing Inbound/Outbound — Sheet: {sheet_name}")
        ws = wb[sheet_name]
        c, o = import_main_sheet(db, ws, cfg["client"], cfg["col_map"], start_row=3)
        grand_containers += c
        grand_outbound   += o

    if "FbaCarriers" in wb.sheetnames:
        section("Importing FBA Carrier Appointments — Sheet: FbaCarriers")
        import_carriers(db, wb["FbaCarriers"], start_row=2)

    if "DOCKING FORM" in wb.sheetnames:
        section("Importing Docking Records — Sheet: DOCKING FORM")
        import_docking(db, wb["DOCKING FORM"], start_row=6)

    db.close()

    section("IMPORT COMPLETE")
    print(f"  Total containers (inbound) : {grand_containers}")
    print(f"  Total outbound records     : {grand_outbound}")
    print(f"\n  Open http://localhost:8000 to see your data.\n")

if __name__ == "__main__":
    main()
