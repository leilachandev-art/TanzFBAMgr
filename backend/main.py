"""
FBA Manager - Flask Backend
Pure Python, no Rust extensions, works on Python 3.14+
"""
from flask import Flask, request, jsonify, send_file, g
from flask_cors import CORS
from sqlalchemy.orm import Session
from sqlalchemy import func
from functools import wraps
import jwt
import bcrypt
import os
import io
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database import engine, SessionLocal, Base
import models

# ── Init ──────────────────────────────────────────────────────────────────────
Base.metadata.create_all(bind=engine)

app = Flask(__name__, static_folder=None)
CORS(app, resources={r"/api/*": {"origins": "*"}})

SECRET_KEY = "fba-mgr-secret-key-2025-change-in-production"
TOKEN_EXPIRE_HOURS = 8

# ── Helpers ───────────────────────────────────────────────────────────────────
def get_db():
    return SessionLocal()

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def check_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode(), hashed.encode())
    except Exception:
        return False

def make_token(username: str) -> str:
    payload = {
        "sub": username,
        "exp": datetime.utcnow() + timedelta(hours=TOKEN_EXPIRE_HOURS),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")

def decode_token(token: str):
    return jwt.decode(token, SECRET_KEY, algorithms=["HS256"])

def dt_str(dt):
    return dt.isoformat() if dt else None

def ser_user(u):
    return {"id": u.id, "username": u.username, "full_name": u.full_name,
            "email": u.email, "role": u.role, "is_active": u.is_active,
            "created_at": dt_str(u.created_at)}

def ser_client(c):
    return {"id": c.id, "code": c.code, "name": c.name, "contact": c.contact,
            "phone": c.phone, "email": c.email, "note": c.note,
            "is_active": c.is_active, "created_at": dt_str(c.created_at)}

def ser_container(c, db):
    agg = db.query(
        func.coalesce(func.sum(models.OutboundRecord.ctns_out), 0),
        func.coalesce(func.sum(models.OutboundRecord.skids_out), 0),
    ).filter(models.OutboundRecord.container_id == c.id).first()
    ctns_out, skids_out = agg[0], agg[1]
    return {
        "id": c.id, "client_code": c.client_code, "date_in": dt_str(c.date_in),
        "container_type": c.container_type, "container_no": c.container_no,
        "mark": c.mark, "inspection": c.inspection, "ctns_in": c.ctns_in,
        "skids_in": c.skids_in, "sku": c.sku, "destination": c.destination,
        "note": c.note, "status": c.status, "created_by": c.created_by,
        "created_at": dt_str(c.created_at),
        "ctns_out_total": ctns_out, "skids_out_total": skids_out,
        "ctns_balance": max(0, c.ctns_in - ctns_out),
        "skids_balance": max(0, c.skids_in - skids_out),
    }

def ser_outbound(r):
    container = r.container
    return {
        "id": r.id, "container_id": r.container_id,
        "container_no": container.container_no if container else None,
        "client_code": container.client_code if container else None,
        "date_out": dt_str(r.date_out), "ctns_out": r.ctns_out,
        "skids_out": r.skids_out, "carrier": r.carrier,
        "destination": r.destination, "isa": r.isa, "note": r.note,
        "pod": r.pod, "wait_time": r.wait_time, "time_in": r.time_in,
        "time_out": r.time_out, "total_pallets": r.total_pallets,
        "created_by": r.created_by, "created_at": dt_str(r.created_at),
    }

def ser_appointment(a):
    return {
        "id": a.id, "client_code": a.client_code,
        "appointment_id": a.appointment_id, "trailer_number": a.trailer_number,
        "reference_code": a.reference_code, "destination_fc": a.destination_fc,
        "isa": a.isa, "carrier": a.carrier, "status": a.status,
        "requested_delivery_date": a.requested_delivery_date,
        "earliest_arrival_time": a.earliest_arrival_time,
        "scheduled_time": a.scheduled_time, "arrival_time": a.arrival_time,
        "checkin_time": a.checkin_time, "unloaded_time": a.unloaded_time,
        "closed_time": a.closed_time, "creation_time": a.creation_time,
        "note": a.note, "created_by": a.created_by,
        "created_at": dt_str(a.created_at),
    }

def ser_docking(d):
    return {
        "id": d.id, "date_in": dt_str(d.date_in), "dock_no": d.dock_no,
        "container_id": d.container_id, "container_no": d.container_no,
        "carrier": d.carrier, "client_code": d.client_code, "status": d.status,
        "date_out": dt_str(d.date_out), "note": d.note,
        "created_by": d.created_by, "created_at": dt_str(d.created_at),
    }

# ── Auth decorator ────────────────────────────────────────────────────────────
def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get("Authorization", "")
        if not auth_header.startswith("Bearer "):
            return jsonify({"detail": "Missing token"}), 401
        token = auth_header.split(" ", 1)[1]
        try:
            payload = decode_token(token)
            username = payload["sub"]
        except jwt.ExpiredSignatureError:
            return jsonify({"detail": "Token expired"}), 401
        except Exception:
            return jsonify({"detail": "Invalid token"}), 401
        db = get_db()
        user = db.query(models.User).filter(models.User.username == username, models.User.is_active == True).first()
        db.close()
        if not user:
            return jsonify({"detail": "User not found"}), 401
        g.current_user = user
        return f(*args, **kwargs)
    return decorated

def require_admin(f):
    @wraps(f)
    @require_auth
    def decorated(*args, **kwargs):
        if g.current_user.role not in ("admin", "manager"):
            return jsonify({"detail": "Insufficient permissions"}), 403
        return f(*args, **kwargs)
    return decorated

# ── Auth routes ───────────────────────────────────────────────────────────────
@app.post("/api/auth/token")
def login():
    username = request.form.get("username") or request.json.get("username") if request.is_json else request.form.get("username")
    password = request.form.get("password") or (request.json.get("password") if request.is_json else None)
    if not username or not password:
        data = request.get_json(silent=True) or {}
        username = data.get("username", username)
        password = data.get("password", password)
    db = get_db()
    try:
        user = db.query(models.User).filter(models.User.username == username).first()
        if not user or not check_password(password, user.hashed_password):
            return jsonify({"detail": "Incorrect username or password"}), 401
        if not user.is_active:
            return jsonify({"detail": "Account disabled"}), 400
        token = make_token(user.username)
        return jsonify({
            "access_token": token, "token_type": "bearer",
            "user": {"id": user.id, "username": user.username,
                     "full_name": user.full_name, "role": user.role},
        })
    finally:
        db.close()

@app.get("/api/auth/me")
@require_auth
def me():
    u = g.current_user
    return jsonify({"id": u.id, "username": u.username, "full_name": u.full_name,
                    "role": u.role, "email": u.email})

# ── Health ────────────────────────────────────────────────────────────────────
@app.get("/api/health")
def health():
    return jsonify({"status": "ok", "version": "1.0.0"})

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.get("/api/dashboard")
@require_auth
def dashboard():
    db = get_db()
    try:
        total = db.query(models.Container).count()
        active = db.query(models.Container).filter(models.Container.status == "In Storage").count()
        partial = db.query(models.Container).filter(models.Container.status == "Partially Out").count()
        completed = db.query(models.Container).filter(models.Container.status == "Completed").count()
        total_ctns_in = db.query(func.coalesce(func.sum(models.Container.ctns_in), 0)).scalar()
        total_ctns_out = db.query(func.coalesce(func.sum(models.OutboundRecord.ctns_out), 0)).scalar()
        pending_appts = db.query(models.FbaAppointment).filter(
            models.FbaAppointment.status.in_(["PENDING", "ARRIVAL_SCHEDULED"])).count()
        active_docks = db.query(models.DockingRecord).filter(
            models.DockingRecord.status.in_(["PENDING", "FULL"])).count()
        recent_in = db.query(models.Container).order_by(models.Container.created_at.desc()).limit(5).all()
        recent_out = db.query(models.OutboundRecord).order_by(models.OutboundRecord.created_at.desc()).limit(5).all()
        return jsonify({
            "total_containers": total, "active_containers": active,
            "partial_containers": partial, "completed_containers": completed,
            "total_ctns_in": total_ctns_in, "total_ctns_out": total_ctns_out,
            "ctns_balance": max(0, total_ctns_in - total_ctns_out),
            "pending_appointments": pending_appts, "active_docks": active_docks,
            "recent_inbound": [
                {"id": c.id, "container_no": c.container_no, "client_code": c.client_code,
                 "ctns_in": c.ctns_in, "destination": c.destination, "status": c.status,
                 "date_in": dt_str(c.date_in)} for c in recent_in
            ],
            "recent_outbound": [
                {"id": r.id, "container_no": r.container.container_no if r.container else "",
                 "ctns_out": r.ctns_out, "carrier": r.carrier,
                 "destination": r.destination, "date_out": dt_str(r.date_out)}
                for r in recent_out
            ],
        })
    finally:
        db.close()

# ── Users ─────────────────────────────────────────────────────────────────────
@app.get("/api/users/")
@require_admin
def list_users():
    db = get_db()
    try:
        return jsonify([ser_user(u) for u in db.query(models.User).all()])
    finally:
        db.close()

@app.post("/api/users/")
@require_admin
def create_user():
    data = request.json
    db = get_db()
    try:
        if db.query(models.User).filter(models.User.username == data["username"]).first():
            return jsonify({"detail": "Username already exists"}), 400
        u = models.User(username=data["username"], full_name=data.get("full_name"),
                        email=data.get("email"), role=data.get("role", "staff"),
                        hashed_password=hash_password(data["password"]))
        db.add(u); db.commit(); db.refresh(u)
        return jsonify(ser_user(u)), 201
    finally:
        db.close()

@app.put("/api/users/<int:uid>")
@require_admin
def update_user(uid):
    data = request.json
    db = get_db()
    try:
        u = db.query(models.User).filter(models.User.id == uid).first()
        if not u: return jsonify({"detail": "Not found"}), 404
        for field in ("full_name", "email", "role", "is_active"):
            if field in data: setattr(u, field, data[field])
        if data.get("password"):
            u.hashed_password = hash_password(data["password"])
        db.commit(); db.refresh(u)
        return jsonify(ser_user(u))
    finally:
        db.close()

@app.delete("/api/users/<int:uid>")
@require_admin
def delete_user(uid):
    db = get_db()
    try:
        u = db.query(models.User).filter(models.User.id == uid).first()
        if not u: return jsonify({"detail": "Not found"}), 404
        if u.id == g.current_user.id: return jsonify({"detail": "Cannot delete yourself"}), 400
        db.delete(u); db.commit()
        return jsonify({"ok": True})
    finally:
        db.close()

# ── Clients ───────────────────────────────────────────────────────────────────
@app.get("/api/clients/")
@require_auth
def list_clients():
    search = request.args.get("search", "")
    db = get_db()
    try:
        q = db.query(models.Client)
        if search:
            q = q.filter(models.Client.code.contains(search) | models.Client.name.contains(search))
        return jsonify([ser_client(c) for c in q.order_by(models.Client.code).all()])
    finally:
        db.close()

@app.post("/api/clients/")
@require_auth
def create_client():
    data = request.json
    db = get_db()
    try:
        if db.query(models.Client).filter(models.Client.code == data["code"]).first():
            return jsonify({"detail": "Client code already exists"}), 400
        c = models.Client(**{k: data.get(k) for k in ("code","name","contact","phone","email","note")})
        db.add(c); db.commit(); db.refresh(c)
        return jsonify(ser_client(c)), 201
    finally:
        db.close()

@app.put("/api/clients/<int:cid>")
@require_auth
def update_client(cid):
    data = request.json
    db = get_db()
    try:
        c = db.query(models.Client).filter(models.Client.id == cid).first()
        if not c: return jsonify({"detail": "Not found"}), 404
        for field in ("name","contact","phone","email","note","is_active"):
            if field in data: setattr(c, field, data[field])
        db.commit(); db.refresh(c)
        return jsonify(ser_client(c))
    finally:
        db.close()

@app.delete("/api/clients/<int:cid>")
@require_admin
def delete_client(cid):
    db = get_db()
    try:
        c = db.query(models.Client).filter(models.Client.id == cid).first()
        if not c: return jsonify({"detail": "Not found"}), 404
        db.delete(c); db.commit()
        return jsonify({"ok": True})
    finally:
        db.close()

# ── Inbound ───────────────────────────────────────────────────────────────────
def parse_dt(s):
    if not s: return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00").replace("+00:00", ""))
    except Exception:
        return None

@app.get("/api/inbound/")
@require_auth
def list_inbound():
    args = request.args
    db = get_db()
    try:
        q = db.query(models.Container)
        if args.get("search"):
            s = args["search"]
            q = q.filter(models.Container.container_no.contains(s) |
                         models.Container.client_code.contains(s) |
                         models.Container.mark.contains(s))
        if args.get("client_code"): q = q.filter(models.Container.client_code == args["client_code"])
        if args.get("status"): q = q.filter(models.Container.status == args["status"])
        if args.get("destination"): q = q.filter(models.Container.destination.contains(args["destination"]))
        rows = q.order_by(models.Container.date_in.desc()).limit(int(args.get("limit", 200))).all()
        return jsonify([ser_container(c, db) for c in rows])
    finally:
        db.close()

@app.post("/api/inbound/")
@require_auth
def create_inbound():
    data = request.json
    db = get_db()
    try:
        if not db.query(models.Client).filter(models.Client.code == data["client_code"]).first():
            db.add(models.Client(code=data["client_code"], name=data["client_code"]))
            db.commit()
        c = models.Container(
            client_code=data["client_code"], date_in=parse_dt(data.get("date_in")),
            container_type=data.get("container_type", "container"),
            container_no=data.get("container_no"), mark=data.get("mark"),
            inspection=data.get("inspection"), ctns_in=data.get("ctns_in", 0),
            skids_in=data.get("skids_in", 0), sku=data.get("sku"),
            destination=data.get("destination"), note=data.get("note"),
            status=data.get("status", "In Storage"), created_by=g.current_user.username,
        )
        db.add(c); db.commit(); db.refresh(c)
        return jsonify(ser_container(c, db)), 201
    finally:
        db.close()

@app.get("/api/inbound/<int:cid>")
@require_auth
def get_inbound(cid):
    db = get_db()
    try:
        c = db.query(models.Container).filter(models.Container.id == cid).first()
        if not c: return jsonify({"detail": "Not found"}), 404
        return jsonify(ser_container(c, db))
    finally:
        db.close()

@app.put("/api/inbound/<int:cid>")
@require_auth
def update_inbound(cid):
    data = request.json
    db = get_db()
    try:
        c = db.query(models.Container).filter(models.Container.id == cid).first()
        if not c: return jsonify({"detail": "Not found"}), 404
        fields = ("client_code","container_type","container_no","mark","inspection",
                  "ctns_in","skids_in","sku","destination","note","status")
        for f in fields:
            if f in data: setattr(c, f, data[f])
        if "date_in" in data: c.date_in = parse_dt(data["date_in"])
        db.commit(); db.refresh(c)
        return jsonify(ser_container(c, db))
    finally:
        db.close()

@app.delete("/api/inbound/<int:cid>")
@require_admin
def delete_inbound(cid):
    db = get_db()
    try:
        c = db.query(models.Container).filter(models.Container.id == cid).first()
        if not c: return jsonify({"detail": "Not found"}), 404
        db.delete(c); db.commit()
        return jsonify({"ok": True})
    finally:
        db.close()

# ── Outbound ──────────────────────────────────────────────────────────────────
@app.get("/api/outbound/")
@require_auth
def list_outbound():
    args = request.args
    db = get_db()
    try:
        q = db.query(models.OutboundRecord)
        if args.get("container_id"): q = q.filter(models.OutboundRecord.container_id == int(args["container_id"]))
        if args.get("carrier"): q = q.filter(models.OutboundRecord.carrier.contains(args["carrier"]))
        if args.get("destination"): q = q.filter(models.OutboundRecord.destination.contains(args["destination"]))
        if args.get("search"):
            s = args["search"]
            q = q.join(models.Container, isouter=True).filter(
                models.Container.container_no.contains(s) |
                models.OutboundRecord.carrier.contains(s) |
                models.OutboundRecord.destination.contains(s) |
                models.OutboundRecord.pod.contains(s))
        rows = q.order_by(models.OutboundRecord.date_out.desc()).limit(int(args.get("limit", 200))).all()
        return jsonify([ser_outbound(r) for r in rows])
    finally:
        db.close()

@app.post("/api/outbound/")
@require_auth
def create_outbound():
    data = request.json
    db = get_db()
    try:
        container = db.query(models.Container).filter(models.Container.id == data["container_id"]).first()
        if not container: return jsonify({"detail": "Container not found"}), 404
        r = models.OutboundRecord(
            container_id=data["container_id"], date_out=parse_dt(data.get("date_out")),
            ctns_out=data.get("ctns_out", 0), skids_out=data.get("skids_out", 0),
            carrier=data.get("carrier"), destination=data.get("destination"),
            isa=data.get("isa"), note=data.get("note"), pod=data.get("pod"),
            wait_time=data.get("wait_time"), time_in=data.get("time_in"),
            time_out=data.get("time_out"), total_pallets=data.get("total_pallets"),
            created_by=g.current_user.username,
        )
        db.add(r)
        total_out = db.query(func.coalesce(func.sum(models.OutboundRecord.ctns_out), 0)).filter(
            models.OutboundRecord.container_id == container.id).scalar() + data.get("ctns_out", 0)
        container.status = "Completed" if total_out >= container.ctns_in else "Partially Out"
        db.commit(); db.refresh(r)
        return jsonify(ser_outbound(r)), 201
    finally:
        db.close()

@app.put("/api/outbound/<int:rid>")
@require_auth
def update_outbound(rid):
    data = request.json
    db = get_db()
    try:
        r = db.query(models.OutboundRecord).filter(models.OutboundRecord.id == rid).first()
        if not r: return jsonify({"detail": "Not found"}), 404
        fields = ("ctns_out","skids_out","carrier","destination","isa","note","pod",
                  "wait_time","time_in","time_out","total_pallets")
        for f in fields:
            if f in data: setattr(r, f, data[f])
        if "date_out" in data: r.date_out = parse_dt(data["date_out"])
        db.commit(); db.refresh(r)
        return jsonify(ser_outbound(r))
    finally:
        db.close()

@app.delete("/api/outbound/<int:rid>")
@require_admin
def delete_outbound(rid):
    db = get_db()
    try:
        r = db.query(models.OutboundRecord).filter(models.OutboundRecord.id == rid).first()
        if not r: return jsonify({"detail": "Not found"}), 404
        db.delete(r); db.commit()
        return jsonify({"ok": True})
    finally:
        db.close()

# ── Inventory ─────────────────────────────────────────────────────────────────
@app.get("/api/inventory/")
@require_auth
def get_inventory():
    args = request.args
    db = get_db()
    try:
        q = db.query(models.Container).filter(models.Container.status != "Completed")
        if args.get("search"):
            s = args["search"]
            q = q.filter(models.Container.container_no.contains(s) | models.Container.client_code.contains(s))
        if args.get("client_code"): q = q.filter(models.Container.client_code == args["client_code"])
        results = []
        for c in q.order_by(models.Container.date_in.desc()).all():
            agg = db.query(func.coalesce(func.sum(models.OutboundRecord.ctns_out), 0),
                           func.coalesce(func.sum(models.OutboundRecord.skids_out), 0)
                           ).filter(models.OutboundRecord.container_id == c.id).first()
            results.append({
                "id": c.id, "client_code": c.client_code, "container_no": c.container_no,
                "container_type": c.container_type, "date_in": dt_str(c.date_in),
                "destination": c.destination, "sku": c.sku, "mark": c.mark,
                "ctns_in": c.ctns_in, "skids_in": c.skids_in,
                "ctns_out": agg[0], "skids_out": agg[1],
                "ctns_balance": max(0, c.ctns_in - agg[0]),
                "skids_balance": max(0, c.skids_in - agg[1]),
                "status": c.status,
            })
        return jsonify(results)
    finally:
        db.close()

@app.get("/api/inventory/summary")
@require_auth
def inventory_summary():
    db = get_db()
    try:
        ti = db.query(func.coalesce(func.sum(models.Container.ctns_in), 0)).scalar()
        to_ = db.query(func.coalesce(func.sum(models.OutboundRecord.ctns_out), 0)).scalar()
        si = db.query(func.coalesce(func.sum(models.Container.skids_in), 0)).scalar()
        so_ = db.query(func.coalesce(func.sum(models.OutboundRecord.skids_out), 0)).scalar()
        return jsonify({
            "total_ctns_in": ti, "total_ctns_out": to_, "total_ctns_balance": max(0, ti - to_),
            "total_skids_in": si, "total_skids_out": so_, "total_skids_balance": max(0, si - so_),
            "active_containers": db.query(models.Container).filter(models.Container.status == "In Storage").count(),
            "partial_containers": db.query(models.Container).filter(models.Container.status == "Partially Out").count(),
            "completed_containers": db.query(models.Container).filter(models.Container.status == "Completed").count(),
        })
    finally:
        db.close()

# ── FBA Appointments ──────────────────────────────────────────────────────────
@app.get("/api/carriers/")
@require_auth
def list_carriers():
    args = request.args
    db = get_db()
    try:
        q = db.query(models.FbaAppointment)
        if args.get("search"):
            s = args["search"]
            q = q.filter(models.FbaAppointment.appointment_id.contains(s) |
                         models.FbaAppointment.reference_code.contains(s) |
                         models.FbaAppointment.trailer_number.contains(s) |
                         models.FbaAppointment.carrier.contains(s))
        if args.get("client_code"): q = q.filter(models.FbaAppointment.client_code == args["client_code"])
        if args.get("status"): q = q.filter(models.FbaAppointment.status == args["status"])
        if args.get("destination_fc"): q = q.filter(models.FbaAppointment.destination_fc.contains(args["destination_fc"]))
        rows = q.order_by(models.FbaAppointment.id.desc()).limit(int(args.get("limit", 300))).all()
        return jsonify([ser_appointment(a) for a in rows])
    finally:
        db.close()

@app.post("/api/carriers/")
@require_auth
def create_carrier():
    data = request.json
    db = get_db()
    try:
        fields = ("client_code","appointment_id","trailer_number","reference_code","destination_fc",
                  "isa","carrier","status","requested_delivery_date","earliest_arrival_time",
                  "scheduled_time","arrival_time","checkin_time","unloaded_time","closed_time",
                  "creation_time","note")
        a = models.FbaAppointment(**{f: data.get(f) for f in fields}, created_by=g.current_user.username)
        if not a.status: a.status = "PENDING"
        db.add(a); db.commit(); db.refresh(a)
        return jsonify(ser_appointment(a)), 201
    finally:
        db.close()

@app.put("/api/carriers/<int:aid>")
@require_auth
def update_carrier(aid):
    data = request.json
    db = get_db()
    try:
        a = db.query(models.FbaAppointment).filter(models.FbaAppointment.id == aid).first()
        if not a: return jsonify({"detail": "Not found"}), 404
        for f in ("client_code","appointment_id","trailer_number","reference_code","destination_fc",
                  "isa","carrier","status","requested_delivery_date","earliest_arrival_time",
                  "scheduled_time","arrival_time","checkin_time","unloaded_time","closed_time",
                  "creation_time","note"):
            if f in data: setattr(a, f, data[f])
        db.commit(); db.refresh(a)
        return jsonify(ser_appointment(a))
    finally:
        db.close()

@app.delete("/api/carriers/<int:aid>")
@require_admin
def delete_carrier(aid):
    db = get_db()
    try:
        a = db.query(models.FbaAppointment).filter(models.FbaAppointment.id == aid).first()
        if not a: return jsonify({"detail": "Not found"}), 404
        db.delete(a); db.commit()
        return jsonify({"ok": True})
    finally:
        db.close()

# ── Docking ───────────────────────────────────────────────────────────────────
@app.get("/api/docking/")
@require_auth
def list_docking():
    args = request.args
    db = get_db()
    try:
        q = db.query(models.DockingRecord)
        if args.get("search"):
            s = args["search"]
            q = q.filter(models.DockingRecord.container_no.contains(s) |
                         models.DockingRecord.client_code.contains(s) |
                         models.DockingRecord.carrier.contains(s))
        if args.get("status"): q = q.filter(models.DockingRecord.status == args["status"])
        rows = q.order_by(models.DockingRecord.date_in.desc()).limit(int(args.get("limit", 200))).all()
        return jsonify([ser_docking(d) for d in rows])
    finally:
        db.close()

@app.post("/api/docking/")
@require_auth
def create_docking():
    data = request.json
    db = get_db()
    try:
        d = models.DockingRecord(
            date_in=parse_dt(data.get("date_in")), dock_no=data.get("dock_no"),
            container_id=data.get("container_id"), container_no=data.get("container_no"),
            carrier=data.get("carrier"), client_code=data.get("client_code"),
            status=data.get("status", "PENDING"), date_out=parse_dt(data.get("date_out")),
            note=data.get("note"), created_by=g.current_user.username,
        )
        db.add(d); db.commit(); db.refresh(d)
        return jsonify(ser_docking(d)), 201
    finally:
        db.close()

@app.put("/api/docking/<int:did>")
@require_auth
def update_docking(did):
    data = request.json
    db = get_db()
    try:
        d = db.query(models.DockingRecord).filter(models.DockingRecord.id == did).first()
        if not d: return jsonify({"detail": "Not found"}), 404
        for f in ("dock_no","container_no","carrier","client_code","status","note"):
            if f in data: setattr(d, f, data[f])
        if "date_in" in data: d.date_in = parse_dt(data["date_in"])
        if "date_out" in data: d.date_out = parse_dt(data["date_out"])
        db.commit(); db.refresh(d)
        return jsonify(ser_docking(d))
    finally:
        db.close()

@app.delete("/api/docking/<int:did>")
@require_admin
def delete_docking(did):
    db = get_db()
    try:
        d = db.query(models.DockingRecord).filter(models.DockingRecord.id == did).first()
        if not d: return jsonify({"detail": "Not found"}), 404
        db.delete(d); db.commit()
        return jsonify({"ok": True})
    finally:
        db.close()

# ── Excel Export ──────────────────────────────────────────────────────────────
def _hdr_style(cell, color="1F4E79"):
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center")

@app.get("/api/excel/export/inbound")
@require_auth
def export_inbound():
    db = get_db()
    try:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Inbound"
        hdrs = ["ID","Client","Date In","Type","Container#","Mark","Inspection",
                "CTNS In","SKIDs In","SKU","Destination","Status","Note","Created By"]
        for i, h in enumerate(hdrs, 1): _hdr_style(ws.cell(1, i, h))
        for ri, c in enumerate(db.query(models.Container).order_by(models.Container.date_in.desc()).all(), 2):
            for ci, v in enumerate([c.id, c.client_code,
                c.date_in.strftime("%Y-%m-%d") if c.date_in else "",
                c.container_type, c.container_no, c.mark, c.inspection,
                c.ctns_in, c.skids_in, c.sku, c.destination, c.status, c.note, c.created_by], 1):
                ws.cell(ri, ci, v)
        for i in range(1, len(hdrs)+1): ws.column_dimensions[get_column_letter(i)].width = 16
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"FBA_Inbound_{datetime.now().strftime('%Y%m%d')}.xlsx")
    finally:
        db.close()

@app.get("/api/excel/export/outbound")
@require_auth
def export_outbound():
    db = get_db()
    try:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Outbound"
        hdrs = ["ID","Container#","Client","Date Out","CTNS Out","SKIDs Out",
                "Carrier","Destination FC","ISA","POD","Wait Time","Time In","Time Out","Pallets","Note"]
        for i, h in enumerate(hdrs, 1): _hdr_style(ws.cell(1, i, h))
        for ri, r in enumerate(db.query(models.OutboundRecord).order_by(models.OutboundRecord.date_out.desc()).all(), 2):
            c = r.container
            for ci, v in enumerate([r.id, c.container_no if c else "", c.client_code if c else "",
                r.date_out.strftime("%Y-%m-%d") if r.date_out else "",
                r.ctns_out, r.skids_out, r.carrier, r.destination, r.isa,
                r.pod, r.wait_time, r.time_in, r.time_out, r.total_pallets, r.note], 1):
                ws.cell(ri, ci, v)
        for i in range(1, len(hdrs)+1): ws.column_dimensions[get_column_letter(i)].width = 16
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"FBA_Outbound_{datetime.now().strftime('%Y%m%d')}.xlsx")
    finally:
        db.close()

@app.get("/api/excel/export/inventory")
@require_auth
def export_inventory():
    db = get_db()
    try:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Inventory"
        hdrs = ["Container#","Client","Date In","Destination","CTNS In","CTNS Out","CTNS Balance",
                "SKIDs In","SKIDs Out","SKIDs Balance","Status"]
        for i, h in enumerate(hdrs, 1): _hdr_style(ws.cell(1, i, h))
        containers = db.query(models.Container).filter(models.Container.status != "Completed").order_by(models.Container.date_in.desc()).all()
        for ri, c in enumerate(containers, 2):
            agg = db.query(func.coalesce(func.sum(models.OutboundRecord.ctns_out), 0),
                           func.coalesce(func.sum(models.OutboundRecord.skids_out), 0)
                           ).filter(models.OutboundRecord.container_id == c.id).first()
            for ci, v in enumerate([c.container_no, c.client_code,
                c.date_in.strftime("%Y-%m-%d") if c.date_in else "",
                c.destination, c.ctns_in, agg[0], max(0, c.ctns_in - agg[0]),
                c.skids_in, agg[1], max(0, c.skids_in - agg[1]), c.status], 1):
                ws.cell(ri, ci, v)
        for i in range(1, len(hdrs)+1): ws.column_dimensions[get_column_letter(i)].width = 16
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"FBA_Inventory_{datetime.now().strftime('%Y%m%d')}.xlsx")
    finally:
        db.close()

@app.get("/api/excel/export/carriers")
@require_auth
def export_carriers():
    db = get_db()
    try:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Appointments"
        hdrs = ["Client","Appointment ID","Trailer#","Ref Code","Dest FC","ISA","Carrier","Status",
                "Requested Delivery","Scheduled Time","Arrival","Check In","Unloaded","Closed","Note"]
        for i, h in enumerate(hdrs, 1): _hdr_style(ws.cell(1, i, h))
        for ri, a in enumerate(db.query(models.FbaAppointment).order_by(models.FbaAppointment.id.desc()).all(), 2):
            for ci, v in enumerate([a.client_code, a.appointment_id, a.trailer_number,
                a.reference_code, a.destination_fc, a.isa, a.carrier, a.status,
                a.requested_delivery_date, a.scheduled_time, a.arrival_time,
                a.checkin_time, a.unloaded_time, a.closed_time, a.note], 1):
                ws.cell(ri, ci, v)
        for i in range(1, len(hdrs)+1): ws.column_dimensions[get_column_letter(i)].width = 18
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"FBA_Carriers_{datetime.now().strftime('%Y%m%d')}.xlsx")
    finally:
        db.close()

# ── Serve React frontend ───────────────────────────────────────────────────────
FRONTEND_DIST = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "frontend", "dist")

@app.route("/assets/<path:filename>")
def serve_assets(filename):
    return send_file(os.path.join(FRONTEND_DIST, "assets", filename))

@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_react(path):
    index = os.path.join(FRONTEND_DIST, "index.html")
    if os.path.exists(index):
        return send_file(index)
    return jsonify({"message": "FBA Manager API running. Run install.bat to build frontend."}), 200

# ── Seed & run ────────────────────────────────────────────────────────────────
def seed():
    db = SessionLocal()
    try:
        if not db.query(models.User).filter(models.User.username == "admin").first():
            db.add(models.User(username="admin", full_name="Administrator",
                               hashed_password=hash_password("admin123"), role="admin", is_active=True))
            for code in ["3003","3002","3014","3008","3033","2025","2039","Voltz","MIDEA"]:
                if not db.query(models.Client).filter(models.Client.code == code).first():
                    db.add(models.Client(code=code, name=code))
            db.commit()
            print("✅ Admin seeded: admin / admin123")
    finally:
        db.close()

seed()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    debug = os.environ.get("FLASK_DEBUG", "true").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
