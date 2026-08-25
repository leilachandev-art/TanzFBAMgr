"""Excel import / export endpoints"""
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
from database import get_db
import models, auth

router = APIRouter(prefix="/api/excel", tags=["excel"])


@router.get("/export/inbound")
def export_inbound(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inbound"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["ID", "Client", "Date In", "Type", "Container#", "Mark", "Inspection",
               "CTNS In", "SKIDS In", "SKU", "Destination", "Status", "Note", "Created By", "Created At"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    containers = db.query(models.Container).order_by(models.Container.date_in.desc()).all()
    for row_idx, c in enumerate(containers, 2):
        ws.cell(row=row_idx, column=1, value=c.id)
        ws.cell(row=row_idx, column=2, value=c.client_code)
        ws.cell(row=row_idx, column=3, value=c.date_in.strftime("%Y-%m-%d") if c.date_in else "")
        ws.cell(row=row_idx, column=4, value=c.container_type)
        ws.cell(row=row_idx, column=5, value=c.container_no)
        ws.cell(row=row_idx, column=6, value=c.mark)
        ws.cell(row=row_idx, column=7, value=c.inspection)
        ws.cell(row=row_idx, column=8, value=c.ctns_in)
        ws.cell(row=row_idx, column=9, value=c.skids_in)
        ws.cell(row=row_idx, column=10, value=c.sku)
        ws.cell(row=row_idx, column=11, value=c.destination)
        ws.cell(row=row_idx, column=12, value=c.status)
        ws.cell(row=row_idx, column=13, value=c.note)
        ws.cell(row=row_idx, column=14, value=c.created_by)
        ws.cell(row=row_idx, column=15, value=c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"FBA_Inbound_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/outbound")
def export_outbound(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outbound"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["ID", "Container#", "Client", "Date Out", "CTNS Out", "SKIDS Out",
               "Carrier", "Destination FC", "ISA", "POD", "Wait Time", "Time In", "Time Out", "Pallets", "Note"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    records = db.query(models.OutboundRecord).order_by(models.OutboundRecord.date_out.desc()).all()
    for row_idx, r in enumerate(records, 2):
        container = r.container
        ws.cell(row=row_idx, column=1, value=r.id)
        ws.cell(row=row_idx, column=2, value=container.container_no if container else "")
        ws.cell(row=row_idx, column=3, value=container.client_code if container else "")
        ws.cell(row=row_idx, column=4, value=r.date_out.strftime("%Y-%m-%d") if r.date_out else "")
        ws.cell(row=row_idx, column=5, value=r.ctns_out)
        ws.cell(row=row_idx, column=6, value=r.skids_out)
        ws.cell(row=row_idx, column=7, value=r.carrier)
        ws.cell(row=row_idx, column=8, value=r.destination)
        ws.cell(row=row_idx, column=9, value=r.isa)
        ws.cell(row=row_idx, column=10, value=r.pod)
        ws.cell(row=row_idx, column=11, value=r.wait_time)
        ws.cell(row=row_idx, column=12, value=r.time_in)
        ws.cell(row=row_idx, column=13, value=r.time_out)
        ws.cell(row=row_idx, column=14, value=r.total_pallets)
        ws.cell(row=row_idx, column=15, value=r.note)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"FBA_Outbound_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/inventory")
def export_inventory(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Container#", "Client", "Date In", "Destination", "CTNS In", "CTNS Out",
               "CTNS Balance", "SKIDS In", "SKIDS Out", "SKIDS Balance", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    containers = db.query(models.Container).filter(
        models.Container.status != "Completed").order_by(models.Container.date_in.desc()).all()
    for row_idx, c in enumerate(containers, 2):
        agg = db.query(
            func.coalesce(func.sum(models.OutboundRecord.ctns_out), 0),
            func.coalesce(func.sum(models.OutboundRecord.skids_out), 0),
        ).filter(models.OutboundRecord.container_id == c.id).first()
        ws.cell(row=row_idx, column=1, value=c.container_no)
        ws.cell(row=row_idx, column=2, value=c.client_code)
        ws.cell(row=row_idx, column=3, value=c.date_in.strftime("%Y-%m-%d") if c.date_in else "")
        ws.cell(row=row_idx, column=4, value=c.destination)
        ws.cell(row=row_idx, column=5, value=c.ctns_in)
        ws.cell(row=row_idx, column=6, value=agg[0])
        ws.cell(row=row_idx, column=7, value=max(0, c.ctns_in - agg[0]))
        ws.cell(row=row_idx, column=8, value=c.skids_in)
        ws.cell(row=row_idx, column=9, value=agg[1])
        ws.cell(row=row_idx, column=10, value=max(0, c.skids_in - agg[1]))
        ws.cell(row=row_idx, column=11, value=c.status)
        # Color balance cells
        if max(0, c.ctns_in - agg[0]) > 0:
            ws.cell(row=row_idx, column=7).fill = PatternFill("solid", fgColor="FFF2CC")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"FBA_Inventory_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/carriers")
def export_carriers(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FBA Appointments"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Client", "Appointment ID", "Trailer#", "Ref Code", "Dest FC", "ISA",
               "Carrier", "Status", "Requested Delivery", "Scheduled Time", "Arrival Time",
               "Check In", "Unloaded", "Closed", "Note"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    appts = db.query(models.FbaAppointment).order_by(models.FbaAppointment.id.desc()).all()
    for row_idx, a in enumerate(appts, 2):
        ws.cell(row=row_idx, column=1, value=a.client_code)
        ws.cell(row=row_idx, column=2, value=a.appointment_id)
        ws.cell(row=row_idx, column=3, value=a.trailer_number)
        ws.cell(row=row_idx, column=4, value=a.reference_code)
        ws.cell(row=row_idx, column=5, value=a.destination_fc)
        ws.cell(row=row_idx, column=6, value=a.isa)
        ws.cell(row=row_idx, column=7, value=a.carrier)
        ws.cell(row=row_idx, column=8, value=a.status)
        ws.cell(row=row_idx, column=9, value=a.requested_delivery_date)
        ws.cell(row=row_idx, column=10, value=a.scheduled_time)
        ws.cell(row=row_idx, column=11, value=a.arrival_time)
        ws.cell(row=row_idx, column=12, value=a.checkin_time)
        ws.cell(row=row_idx, column=13, value=a.unloaded_time)
        ws.cell(row=row_idx, column=14, value=a.closed_time)
        ws.cell(row=row_idx, column=15, value=a.note)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"FBA_Carriers_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename={filename}"})
